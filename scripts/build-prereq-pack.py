"""
Build the RCSA Prerequisites Packs ingested at the start of the demo.

This script generates THREE scenario packs of the same in-scope app
(PAYMENTS-CORE) with materially different risk postures so the demo can show
the AI checkpoints reasoning about distinctly different inputs:

    docs/samples/RCSA-Prerequisites-Pack-PAYMENTS-CORE-2026Q2-CLEAN.xlsx
    docs/samples/RCSA-Prerequisites-Pack-PAYMENTS-CORE-2026Q2-MIXED.xlsx
    docs/samples/RCSA-Prerequisites-Pack-PAYMENTS-CORE-2026Q2-CRITICAL.xlsx

Each pack has the same five tabs (Risk & Control Inventory · Asset & Ownership
· Prior RCSA & Open Issues · Policies & Regulatory · Evidence & Monitoring),
populated with scenario-specific data. The role titles and named individuals
match workflow.js / CLAUDE.md so files feel native to the demo.

Setup (one-time):
    py -m pip install openpyxl

Run:
    py scripts/build-prereq-pack.py
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = REPO_ROOT / "docs" / "samples"

CYCLE_BASE_ID = "RCSA-2026-Q2-PAYC-001"
IN_SCOPE_APP = "PAYMENTS-CORE"
GENERATED = date(2026, 5, 5).isoformat()
OWNER = "M. Chen, Cyber RCSA Lead, 1LOD"


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF1F3A5F")
COVER_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
COVER_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FF1F3A5F")
SCENARIO_BANNER_FONT = Font(name="Calibri", size=12, bold=True, italic=True, color="FFB3321A")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, min_w=10, max_w=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 2))


# Common policies and assets — these are largely scenario-independent.
COMMON_POLICIES = [
    ("NIST CSF 2.0",  "PR.AC-1 Identity & access management",         "PAYMENTS-CORE-API, PAYMENTS-AUTH-SVC", "POL-IAM-001",    "2026-01-15"),
    ("NIST CSF 2.0",  "DE.AE-3 Event data aggregation",                "PAYMENTS-LOG-PIPELINE",                 "POL-LOG-004",    "2026-02-02"),
    ("NIST CSF 2.0",  "PR.DS-1 Data-at-rest protection",               "PAYMENTS-CORE-DB, PAYMENTS-VAULT-SVC", "POL-CRYPTO-002", "2025-11-10"),
    ("NYDFS 500.07",  "Access privileges & management",                "All Tier 1 systems",                    "POL-IAM-001",    "2026-01-15"),
    ("NYDFS 500.14",  "Training and monitoring",                       "Continuous Monitoring Team",            "POL-MON-003",    "2025-12-01"),
    ("NYDFS 500.16",  "Incident response plan",                        "Cyber RCSA Lead, IR Team",              "POL-IR-001",     "2026-03-22"),
    ("SOX ITGC",      "Logical access controls",                       "PAYMENTS-CORE-API, PAYMENTS-CORE-DB",   "POL-SOX-IAM-1",  "2026-01-30"),
    ("SOX ITGC",      "Change management",                             "All in-scope apps",                     "POL-SOX-CM-1",   "2026-01-30"),
    ("PCI-DSS 4.0",   "Req 7 — Restrict access by business need",      "PAYMENTS-CORE-API, PAYMENTS-AUTH-SVC",  "POL-PCI-A7",     "2025-10-05"),
    ("PCI-DSS 4.0",   "Req 10 — Log and monitor all access",           "PAYMENTS-LOG-PIPELINE",                 "POL-PCI-A10",    "2025-10-05"),
    ("PCI-DSS 4.0",   "Req 3 — Protect stored account data",           "PAYMENTS-CORE-DB",                      "POL-PCI-A3",     "2025-10-05"),
    ("Internal",      "Cyber Risk Appetite Statement (Board-approved)","Enterprise-wide",                       "POL-RISK-001",   "2026-02-14"),
]

COMMON_ASSETS = [
    ("CMDB-PAYC-001", "PAYMENTS-CORE-API",      "Tier 1", "L. Ortega",  "Platform Eng — Payments", "Yes", "Yes", "Production"),
    ("CMDB-PAYC-002", "PAYMENTS-CORE-DB",       "Tier 1", "L. Ortega",  "DBA — Payments",           "Yes", "Yes", "Production"),
    ("CMDB-PAYC-003", "PAYMENTS-AUTH-SVC",      "Tier 1", "L. Ortega",  "P. Singh (IAM)",           "Yes", "No",  "Production"),
    ("CMDB-PAYC-004", "PAYMENTS-LOG-PIPELINE",  "Tier 2", "L. Ortega",  "D. Park (Logging Eng)",    "No",  "No",  "Production"),
    ("CMDB-PAYC-005", "PAYMENTS-FRAUD-RULES",   "Tier 2", "L. Ortega",  "Risk Analytics",           "Yes", "No",  "Production"),
    ("CMDB-PAYC-006", "PAYMENTS-RECON-BATCH",   "Tier 2", "L. Ortega",  "Platform Eng — Payments", "No",  "Yes", "Production"),
    ("CMDB-PAYC-007", "PAYMENTS-VAULT-SVC",     "Tier 1", "L. Ortega",  "A. Khan (PKI Eng)",        "Yes", "No",  "Production"),
    ("CMDB-PAYC-008", "PAYMENTS-LEGACY-AUTH",   "Tier 2", "L. Ortega",  "D. Park (Logging Eng)",    "No",  "No",  "Sunset 2026Q4"),
]


# ---------- Scenario data ----------

SCENARIOS = {
    # ====================================================================
    # CLEAN — newly remediated, well-governed posture. AI should find few
    # candidate risks and 2LOD should have minimal challenges to raise.
    # ====================================================================
    "CLEAN": {
        "tagline": "Clean posture — minimal issues, well-governed",
        "risks": [
            ("R-01", "Session timeout consistency across auth paths",          "PAYMENTS-AUTH-SVC",    "CTRL-IAM-09",    "P. Singh (IAM Lead)",          "Low"),
            ("R-02", "API rate-limit configuration drift",                      "PAYMENTS-CORE-API",    "CTRL-API-03",    "Platform Eng — Payments",      "Low"),
            ("R-03", "Vendor SLA monitoring on payment processor",              "PAYMENTS-AUTH-SVC",    "CTRL-VENDOR-05", "Vendor Mgmt Lead",             "Moderate"),
            ("R-04", "DR runbook freshness post-platform updates",              "PAYMENTS-CORE-API",    "CTRL-RES-03",    "Resilience Eng Lead",          "Low"),
            ("R-05", "Quarterly access review automation drift",                "PAYMENTS-CORE-API",    "CTRL-IAM-04",    "P. Singh (IAM Lead)",          "Low"),
        ],
        "prior_issues": [
            ("RCSA-2025-Q4-PAYC", "DR runbook not updated post platform migration",  "Closed",     "2024-12-08", "Resilience Eng Lead",      "CTRL-RES-03",    "No"),
        ],
        "evidence": [
            ("Splunk",        "Auth-path SIEM coverage healthy",                "Stable",            "PAYMENTS-AUTH-SVC",     "2026-05-04"),
            ("Splunk",        "Privileged-session monitoring within baseline",  "Stable",            "PAYMENTS-CORE-API",     "2026-05-04"),
            ("Sentinel",      "API key reuse anomalies — none in last 24mo",    "Stable",            "PAYMENTS-AUTH-SVC",     "2026-04-29"),
            ("KRI Dashboard", "KRI-AUTH-FAIL — auth failure rate",              "Within tolerance",  "PAYMENTS-AUTH-SVC",     "2026-05-03"),
            ("KRI Dashboard", "KRI-IAM-REVIEW — privileged access review",      "Within tolerance",  "PAYMENTS-CORE-API",     "2026-05-01"),
            ("KRI Dashboard", "KRI-PATCH-SLA — critical patch SLA breach rate", "Within tolerance",  "Enterprise-wide",       "2026-05-02"),
            ("Internal Audit","Last finding closed 2025-Q1",                    "No open findings",  "PAYMENTS-CORE-API",     "2025-03-04"),
            ("IR (24mo)",     "No payments-tagged IRs in last 24 months",       "Stable",            "Enterprise-wide",       "2026-05-02"),
            ("Reg Bulletin",  "NYDFS 500 amendments — November 2025 (assessed)","Action complete",   "Enterprise-wide",       "2026-02-15"),
            ("Vuln Mgmt",     "All Critical CVEs patched within SLA",           "Within tolerance",  "PAYMENTS-CORE-API",     "2026-04-30"),
        ],
    },

    # ====================================================================
    # MIXED — current production reality. 10 risks, recurring gaps,
    # mixed control test results. Equivalent to the v9.x default pack.
    # ====================================================================
    "MIXED": {
        "tagline": "Mixed posture — recurring gaps, real-world middle ground",
        "risks": [
            ("R-01", "Privileged access drift on payment processing servers",     "PAYMENTS-CORE-API",    "CTRL-IAM-04",     "P. Singh (IAM Lead)",          "High"),
            ("R-02", "Third-party API key rotation gap",                          "PAYMENTS-AUTH-SVC",    "CTRL-CRYPTO-07",  "A. Khan (PKI Eng Lead)",       "High"),
            ("R-03", "Insufficient logging on legacy auth path",                  "PAYMENTS-LOG-PIPELINE","CTRL-LOG-11",     "D. Park (Logging Eng Lead)",   "High"),
            ("R-04", "Unpatched library in payment gateway component",            "PAYMENTS-CORE-API",    "CTRL-VULN-02",    "Vuln Mgmt Lead",               "High"),
            ("R-05", "Data classification drift on customer PII tables",          "PAYMENTS-CORE-DB",     "CTRL-DATA-09",    "Data Gov Lead",                "Moderate"),
            ("R-06", "Backup encryption key custody — control owner ambiguity",   "PAYMENTS-CORE-DB",     "CTRL-CRYPTO-12",  "PKI Eng Lead / Backup Eng",    "Moderate"),
            ("R-07", "Session timeout policy not enforced uniformly",             "PAYMENTS-AUTH-SVC",    "CTRL-IAM-09",     "P. Singh (IAM Lead)",          "Moderate"),
            ("R-08", "Insider threat — privileged user activity not reviewed",    "PAYMENTS-CORE-API",    "CTRL-IAM-04",     "P. Singh (IAM Lead)",          "Moderate"),
            ("R-09", "DR / failover configuration drift",                          "PAYMENTS-CORE-API",    "CTRL-RES-03",     "Resilience Eng Lead",          "Low"),
            ("R-10", "Vendor SLA monitoring gap on payment processor",            "PAYMENTS-AUTH-SVC",    "CTRL-VENDOR-05",  "Vendor Mgmt Lead",             "Low"),
        ],
        "prior_issues": [
            ("RCSA-2025-Q2-PAYC", "Privileged access review cadence not met (2 quarters)",   "Open",       "2025-07-12", "P. Singh (IAM Lead)",        "CTRL-IAM-04",     "Yes"),
            ("RCSA-2025-Q2-PAYC", "Splunk coverage gap on legacy auth path",                  "Open",       "2025-07-12", "D. Park (Logging Eng Lead)", "CTRL-LOG-11",     "Yes"),
            ("RCSA-2025-Q2-PAYC", "API key rotation evidence not retained",                   "In-progress","2025-09-04", "A. Khan (PKI Eng Lead)",     "CTRL-CRYPTO-07",  "Yes"),
            ("RCSA-2024-Q4-PAYC", "DR runbook not updated post platform migration",           "Closed",     "2024-12-08", "Resilience Eng Lead",        "CTRL-RES-03",     "No"),
            ("AUDIT-2024-INT-22", "Third-party API key rotation gap (audit finding)",         "Open",       "2024-11-19", "A. Khan (PKI Eng Lead)",     "CTRL-CRYPTO-07",  "Yes"),
            ("IR-2024-08",        "PII classification drift surfaced in incident review",     "Open",       "2024-08-22", "Data Gov Lead",              "CTRL-DATA-09",    "Yes"),
        ],
        "evidence": [
            ("Splunk",         "Anomalous privileged-session pattern (14d cluster)", "Spiking",       "PAYMENTS-CORE-API",     "2026-05-04"),
            ("Splunk",         "Auth-path coverage gap on legacy endpoint",          "Persistent",    "PAYMENTS-LEGACY-AUTH",  "2026-05-04"),
            ("Sentinel",       "Suspicious API key reuse across geos",                "Recurring",     "PAYMENTS-AUTH-SVC",     "2026-04-29"),
            ("KRI Dashboard",  "KRI-AUTH-FAIL — auth failure rate",                   "Breach (3/4w)", "PAYMENTS-AUTH-SVC",     "2026-05-03"),
            ("KRI Dashboard",  "KRI-IAM-REVIEW — privileged access review cadence",  "Breach",        "PAYMENTS-CORE-API",     "2026-05-01"),
            ("KRI Dashboard",  "KRI-PATCH-SLA — critical patch SLA breach rate",      "Within tol.",   "Enterprise-wide",       "2026-05-02"),
            ("Internal Audit", "Open finding: API key rotation gap",                  "Open since 2024-11", "PAYMENTS-AUTH-SVC", "2026-04-25"),
            ("Internal Audit", "Open finding: privileged review cadence",             "Open since 2025-07", "PAYMENTS-CORE-API", "2026-04-25"),
            ("Internal Audit", "Open finding: PII classification gates",              "Open since 2024-08", "PAYMENTS-CORE-DB",  "2026-04-25"),
            ("Internal Audit", "Closed finding: DR runbook gap",                      "Closed 2025-Q1",     "PAYMENTS-CORE-API", "2025-03-04"),
            ("IR (24mo)",      "IR-2024-08 PII classification drift",                 "1 incident",         "PAYMENTS-CORE-DB",  "2024-08-22"),
            ("IR (24mo)",      "IR-2025-03 Splunk parser silent failure",             "1 incident",         "PAYMENTS-LOG-PIPELINE","2025-03-11"),
            ("IR (24mo)",      "IR-2025-09 Vendor key rotation miss",                 "1 incident",         "PAYMENTS-AUTH-SVC", "2025-09-04"),
            ("Reg Bulletin",   "NYDFS 500 amendments — November 2025",                "New",                "Enterprise-wide",   "2025-11-01"),
            ("Vuln Mgmt",      "CVE-2025-#### in payment-gateway-lib",                "Open",               "PAYMENTS-CORE-API", "2026-04-30"),
        ],
    },

    # ====================================================================
    # CRITICAL — heavy posture. Multiple cycles of unresolved gaps, recent
    # IR, multi-control failures, regulatory bulletin requiring action.
    # AI should surface many risks (including previously missed ones) and
    # 2LOD should have material challenges to raise.
    # ====================================================================
    "CRITICAL": {
        "tagline": "CRITICAL posture — heavy failures, multi-cycle gap recurrence, recent IR",
        "risks": [
            ("R-01", "Privileged access drift on payment processing servers",     "PAYMENTS-CORE-API",    "CTRL-IAM-04",     "P. Singh (IAM Lead)",          "High"),
            ("R-02", "Third-party API key rotation gap",                          "PAYMENTS-AUTH-SVC",    "CTRL-CRYPTO-07",  "A. Khan (PKI Eng Lead)",       "High"),
            ("R-03", "Insufficient logging on legacy auth path",                  "PAYMENTS-LOG-PIPELINE","CTRL-LOG-11",     "D. Park (Logging Eng Lead)",   "High"),
            ("R-04", "Unpatched library in payment gateway component",            "PAYMENTS-CORE-API",    "CTRL-VULN-02",    "Vuln Mgmt Lead",               "High"),
            ("R-05", "Data classification drift on customer PII tables",          "PAYMENTS-CORE-DB",     "CTRL-DATA-09",    "Data Gov Lead",                "High"),
            ("R-06", "Backup encryption key custody — control owner ambiguity",   "PAYMENTS-CORE-DB",     "CTRL-CRYPTO-12",  "PKI Eng Lead / Backup Eng",    "High"),
            ("R-07", "Production SoD violation — devs with prod-write access",    "PAYMENTS-CORE-API",    "CTRL-SOD-02",     "Platform Eng — Payments",      "High"),
            ("R-08", "Privileged credential leak (IR-2026-04, recent)",            "PAYMENTS-AUTH-SVC",    "CTRL-IAM-04",     "P. Singh (IAM Lead)",          "High"),
            ("R-09", "Tokenization control failed PCI test",                       "PAYMENTS-CORE-DB",     "CTRL-PCI-08",     "Risk Analytics",               "High"),
            ("R-10", "Secrets in CI/CD logs (developer self-disclosure)",          "PAYMENTS-CORE-API",    "CTRL-SECRETS-04", "Platform Eng — Payments",      "High"),
            ("R-11", "Insider threat — privileged user activity not reviewed",    "PAYMENTS-CORE-API",    "CTRL-IAM-04",     "P. Singh (IAM Lead)",          "High"),
            ("R-12", "Session timeout policy not enforced uniformly",             "PAYMENTS-AUTH-SVC",    "CTRL-IAM-09",     "P. Singh (IAM Lead)",          "Moderate"),
            ("R-13", "DR / failover configuration drift",                          "PAYMENTS-CORE-API",    "CTRL-RES-03",     "Resilience Eng Lead",          "Moderate"),
            ("R-14", "Vendor SLA monitoring gap on payment processor",            "PAYMENTS-AUTH-SVC",    "CTRL-VENDOR-05",  "Vendor Mgmt Lead",             "Moderate"),
        ],
        "prior_issues": [
            ("RCSA-2025-Q2-PAYC", "Privileged access review cadence not met (3 quarters)",   "Open",       "2025-07-12", "P. Singh (IAM Lead)",        "CTRL-IAM-04",     "Yes"),
            ("RCSA-2025-Q2-PAYC", "Splunk coverage gap on legacy auth path",                  "Open",       "2025-07-12", "D. Park (Logging Eng Lead)", "CTRL-LOG-11",     "Yes"),
            ("RCSA-2025-Q2-PAYC", "API key rotation evidence not retained",                   "Open",       "2025-09-04", "A. Khan (PKI Eng Lead)",     "CTRL-CRYPTO-07",  "Yes"),
            ("RCSA-2025-Q4-PAYC", "Backup key custody owner not assigned",                    "Open",       "2025-12-15", "PKI Eng Lead",               "CTRL-CRYPTO-12",  "Yes"),
            ("RCSA-2025-Q4-PAYC", "Production SoD exception count above tolerance",            "Open",       "2025-12-22", "Platform Eng — Payments",   "CTRL-SOD-02",     "Yes"),
            ("AUDIT-2024-INT-22", "Third-party API key rotation gap (audit finding)",         "Open",       "2024-11-19", "A. Khan (PKI Eng Lead)",     "CTRL-CRYPTO-07",  "Yes"),
            ("AUDIT-2025-EXT-09", "Tokenization PCI control test failed",                     "Open",       "2025-10-04", "Risk Analytics",             "CTRL-PCI-08",     "Yes"),
            ("IR-2024-08",        "PII classification drift surfaced in incident review",     "Open",       "2024-08-22", "Data Gov Lead",              "CTRL-DATA-09",    "Yes"),
            ("IR-2026-04",        "Privileged credential leak — service account exposure",    "Open",       "2026-04-09", "P. Singh (IAM Lead)",        "CTRL-IAM-04",     "Yes"),
        ],
        "evidence": [
            ("Splunk",         "Anomalous privileged-session pattern (multi-week cluster)", "Spiking",          "PAYMENTS-CORE-API",     "2026-05-04"),
            ("Splunk",         "Auth-path coverage gap — production endpoints",            "Persistent",       "PAYMENTS-LEGACY-AUTH",  "2026-05-04"),
            ("Splunk",         "Suspicious data-egress pattern from PII tables",            "New cluster",      "PAYMENTS-CORE-DB",      "2026-05-03"),
            ("Sentinel",       "Suspicious API key reuse across geos",                       "Recurring",        "PAYMENTS-AUTH-SVC",     "2026-04-29"),
            ("Sentinel",       "Service account credential leak detected (IR-2026-04)",     "1 incident",       "PAYMENTS-AUTH-SVC",     "2026-04-09"),
            ("KRI Dashboard",  "KRI-AUTH-FAIL — auth failure rate",                          "Breach (4/4w)",    "PAYMENTS-AUTH-SVC",     "2026-05-03"),
            ("KRI Dashboard",  "KRI-IAM-REVIEW — privileged access review cadence",         "Breach (3 cycles)","PAYMENTS-CORE-API",     "2026-05-01"),
            ("KRI Dashboard",  "KRI-PATCH-SLA — critical patch SLA breach rate",             "Breach",           "PAYMENTS-CORE-API",     "2026-05-02"),
            ("KRI Dashboard",  "KRI-PCI-CTRL — PCI control test pass rate",                  "Breach",           "PAYMENTS-CORE-DB",      "2026-05-01"),
            ("KRI Dashboard",  "KRI-SECRETS — secrets-in-code detection rate",               "Breach",           "PAYMENTS-CORE-API",     "2026-04-30"),
            ("Internal Audit", "Open finding: API key rotation gap (>18mo open)",            "Aging",            "PAYMENTS-AUTH-SVC",     "2026-04-25"),
            ("Internal Audit", "Open finding: privileged review cadence (>10mo)",            "Aging",            "PAYMENTS-CORE-API",     "2026-04-25"),
            ("Internal Audit", "Open finding: PII classification gates (>20mo)",             "Aging",            "PAYMENTS-CORE-DB",      "2026-04-25"),
            ("Internal Audit", "Open finding: tokenization PCI test (>7mo)",                  "Open",             "PAYMENTS-CORE-DB",      "2026-04-25"),
            ("Internal Audit", "Open finding: SoD exception count above tolerance (>5mo)",   "Open",             "PAYMENTS-CORE-API",     "2026-04-25"),
            ("IR (24mo)",      "IR-2024-08 PII classification drift",                         "1 incident",        "PAYMENTS-CORE-DB",      "2024-08-22"),
            ("IR (24mo)",      "IR-2025-03 Splunk parser silent failure",                     "1 incident",        "PAYMENTS-LOG-PIPELINE", "2025-03-11"),
            ("IR (24mo)",      "IR-2025-09 Vendor key rotation miss",                         "1 incident",        "PAYMENTS-AUTH-SVC",     "2025-09-04"),
            ("IR (24mo)",      "IR-2026-04 Privileged credential leak (recent, severity High)","1 incident",      "PAYMENTS-AUTH-SVC",     "2026-04-09"),
            ("Reg Bulletin",   "NYDFS 500 amendments — November 2025 (action overdue)",      "Action overdue",   "Enterprise-wide",       "2025-11-01"),
            ("Reg Bulletin",   "PCI-DSS 4.0 transition assessment — gaps identified",         "Open",             "PAYMENTS-CORE-DB",      "2026-03-15"),
            ("Vuln Mgmt",      "CVE-2025-#### in payment-gateway-lib",                        "Open",             "PAYMENTS-CORE-API",     "2026-04-30"),
            ("Vuln Mgmt",      "CVE-2026-#### in PCI tokenization library",                   "Open",             "PAYMENTS-CORE-DB",      "2026-04-12"),
        ],
    },
}


# ---------- Sheet writers ----------

def write_risks_controls(ws, scenario_name, scenario):
    """Tab 1 doubles as the cover sheet — top rows describe the pack."""
    ws["A1"] = "RCSA Prerequisites Pack"
    ws["A1"].font = COVER_TITLE_FONT
    ws.merge_cells("A1:F1")

    cycle_id = f"{CYCLE_BASE_ID}-{scenario_name}"

    cover_rows = [
        ("Cycle ID",        cycle_id),
        ("Scenario",        f"{scenario_name} — {scenario['tagline']}"),
        ("In-scope app",    IN_SCOPE_APP),
        ("Generated",       GENERATED),
        ("Owner",           OWNER),
        ("Pack contents",   "5 tabs — Risk & Control Inventory · Asset & Ownership · Prior RCSA & Open Issues · Policies & Regulatory · Evidence & Monitoring"),
    ]
    for i, (label, value) in enumerate(cover_rows, start=2):
        ws.cell(row=i, column=1, value=label).font = COVER_LABEL_FONT
        cell = ws.cell(row=i, column=2, value=value)
        cell.alignment = LEFT
        if label == "Scenario":
            cell.font = SCENARIO_BANNER_FONT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    header_row = 2 + len(cover_rows) + 1
    cols = ["Risk ID", "Risk", "Linked Application", "Mapped Controls", "Control Owner", "Inherent Rating"]
    for c, name in enumerate(cols, start=1):
        ws.cell(row=header_row, column=c, value=name)
    style_header(ws, header_row, len(cols))

    for r_idx, row in enumerate(scenario["risks"], start=header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def write_assets(ws, scenario_name, scenario):
    cols = ["App ID (CMDB)", "App Name", "Tier", "Application Owner", "Tech Owner", "PCI", "SOX", "Lifecycle"]
    for c, name in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=name)
    style_header(ws, 1, len(cols))

    for r_idx, row in enumerate(COMMON_ASSETS, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def write_prior_issues(ws, scenario_name, scenario):
    cols = ["Cycle ID", "Finding / Gap", "Status", "Open Date", "Owner", "Linked Control", "Carry-forward"]
    for c, name in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=name)
    style_header(ws, 1, len(cols))

    for r_idx, row in enumerate(scenario["prior_issues"], start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def write_policies(ws, scenario_name, scenario):
    cols = ["Standard", "Section", "Applicability", "Internal Policy ID", "Last Reviewed"]
    for c, name in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=name)
    style_header(ws, 1, len(cols))

    for r_idx, row in enumerate(COMMON_POLICIES, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def write_evidence(ws, scenario_name, scenario):
    cols = ["Source", "Signal / Finding", "24-mo Trend", "Linked Asset", "Last Updated"]
    for c, name in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=name)
    style_header(ws, 1, len(cols))

    for r_idx, row in enumerate(scenario["evidence"], start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


# ---------- Build one pack ----------

def build_pack(scenario_name, scenario):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "1. Risk & Control Inventory"
    write_risks_controls(ws1, scenario_name, scenario)

    ws2 = wb.create_sheet("2. Asset & Ownership")
    write_assets(ws2, scenario_name, scenario)

    ws3 = wb.create_sheet("3. Prior RCSA & Open Issues")
    write_prior_issues(ws3, scenario_name, scenario)

    ws4 = wb.create_sheet("4. Policies & Regulatory")
    write_policies(ws4, scenario_name, scenario)

    ws5 = wb.create_sheet("5. Evidence & Monitoring")
    write_evidence(ws5, scenario_name, scenario)

    for ws in (ws1, ws2, ws3, ws4, ws5):
        autosize(ws)

    out = OUT_DIR / f"RCSA-Prerequisites-Pack-PAYMENTS-CORE-2026Q2-{scenario_name}.xlsx"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main():
    import shutil

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    mixed_path = None
    for name, scn in SCENARIOS.items():
        out = build_pack(name, scn)
        size = out.stat().st_size
        print(f"Wrote {out.name} ({size:,} bytes · {len(scn['risks'])} risks · "
              f"{len(scn['prior_issues'])} prior issues · {len(scn['evidence'])} evidence rows)")
        if name == "MIXED":
            mixed_path = out

    # Back-compat: also write the legacy single-pack filename as a copy of
    # MIXED. The demo's SAMPLE_PACK_NAME constant + Trigger node attachment
    # both point at this filename.
    if mixed_path is not None:
        legacy = OUT_DIR / "RCSA-Prerequisites-Pack-PAYMENTS-CORE-2026Q2.xlsx"
        shutil.copyfile(mixed_path, legacy)
        print(f"Wrote {legacy.name} (copy of MIXED for back-compat)")


if __name__ == "__main__":
    main()
